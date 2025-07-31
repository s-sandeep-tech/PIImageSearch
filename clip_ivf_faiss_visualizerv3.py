import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import torch
import clip
import numpy as np
from PIL import Image
from tqdm import tqdm
from pathlib import Path
from sklearn.preprocessing import normalize
import subprocess
import faiss
import pickle
import pandas as pd

# ==== CONFIG ====
os.environ["OMP_NUM_THREADS"] = "1"
IMAGE_DIR = "FinalDesignImages"
EXCEL_PATH = "design_metadata.xlsx"  # ✅ Update with your Excel file path
OUTPUT_HTML = "clip_ivf_visualizer.html"
INDEX_PATH = "clip_ivf.index"
PATHS_PKL = "clip_image_paths.pkl"
EMBEDDINGS_NPY = "clip_embeddings.npy"
TOP_K = 5
SIMILARITY_THRESHOLD = 1.0
NLIST = 50
RESTORE_MODE = True  # ✅ Set to True for reload without recomputing


# === CONFIGURATION  XLSX ===
OUTPUT_XLSX = "image_similarity_viewer.xlsx"
MAX_WIDTH = 100   # Max image width in Excel cell (pixels)
MAX_HEIGHT = 100  # Max image height in Excel cell (pixels)

# ==== DEVICE SETUP ====
device = "mps" if torch.backends.mps.is_available() else "cpu"
print(f"Using device: {device}")
model, preprocess = clip.load("ViT-B/32", device=device)

# ==== HELPER: Strip extensions for matching ====
def strip_extension(name):
    return os.path.splitext(os.path.basename(str(name)))[0].lower()

# ==== LOAD METADATA FROM EXCEL (Dynamic Columns) ====
metadata_df = pd.read_excel(EXCEL_PATH)
metadata_df.columns = [str(c) for c in metadata_df.columns]

metadata_lookup = {
    strip_extension(row["ImageName"]): {col: str(val) for col, val in row.items()}
    for _, row in metadata_df.iterrows()
}
print(f"📑 Loaded metadata for {len(metadata_lookup)} images with {len(metadata_df.columns)} fields each.")

# Fields to exclude from HTML display
EXCLUDE_FIELDS = {"Image URL", "Exists", "Cpath", "ImageName"}

# ==== EMBEDDING FUNCTION ====
def get_clip_embedding(img_path):
    try:
        image = Image.open(img_path)
        if image.mode in ["P", "LA"]:
            image = image.convert("RGBA")
        else:
            image = image.convert("RGB")
        image_input = preprocess(image).unsqueeze(0).to(device)
        with torch.no_grad():
            emb = model.encode_image(image_input)
        emb = emb / emb.norm(dim=-1, keepdim=True)
        return emb.detach().cpu().numpy().astype("float32")
    except Exception as e:
        print(f"❌ Failed to embed {img_path}: {e}")
        return None

# ==== RESTORE OR COMPUTE EMBEDDINGS ====
if RESTORE_MODE and os.path.exists(INDEX_PATH) and os.path.exists(PATHS_PKL) and os.path.exists(EMBEDDINGS_NPY):
    print("📥 Loading FAISS index, image paths, and embeddings...")
    index = faiss.read_index(INDEX_PATH)
    with open(PATHS_PKL, "rb") as f:
        valid_paths = pickle.load(f)
    X = np.load(EMBEDDINGS_NPY)
    print(f"✅ Restored {len(valid_paths)} image paths and {X.shape[0]} embeddings from disk.")
else:
    print("🔄 Computing embeddings...")
    image_paths = list(Path(IMAGE_DIR).glob("*.jpg")) + list(Path(IMAGE_DIR).glob("*.jpeg")) + list(Path(IMAGE_DIR).glob("*.png"))
    embeddings = []
    valid_paths = []

    for path in tqdm(image_paths):
        emb = get_clip_embedding(str(path))
        if emb is not None:
            embeddings.append(emb)
            valid_paths.append(path)

    print(f"📦 {len(embeddings)} embeddings computed.")

    X = np.array(embeddings).astype("float32")
    if X.ndim == 3:
        X = X.squeeze()
    X = normalize(X)
    dim = X.shape[1]

    print("📡 Preparing FAISS index...")
    try:
        if len(X) < NLIST * 39:
            raise ValueError("Too few samples for IVF. Using FlatIP.")

        quantizer = faiss.IndexFlatIP(dim)
        index = faiss.IndexIVFFlat(quantizer, dim, NLIST, faiss.METRIC_INNER_PRODUCT)
        print("🔧 Training index...")
        index.train(X)
        index.add(X)
    except Exception as e:
        print(f"⚠️ {e} — fallback to FlatIP...")
        index = faiss.IndexFlatIP(dim)
        index.add(X)

    # ✅ Save for future restore
    faiss.write_index(index, INDEX_PATH)
    with open(PATHS_PKL, "wb") as f:
        pickle.dump(valid_paths, f)
    np.save(EMBEDDINGS_NPY, X)
    print("💾 Saved FAISS index, image paths, and embeddings.")

# ==== RUN SEARCH ====
print("🔍 Running FAISS search...")
D, I = index.search(X, TOP_K + 1)

# ==== HTML GENERATOR ====
def generate_similarity_html(valid_paths, I, D, OUTPUT_HTML, SIMILARITY_THRESHOLD):
    print("🌐 Generating HTML visualizer...")
    skipped = 0

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        # HTML Header and Styling
        f.write(f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Image Similarity Viewer</title>
    <style>
        body {{
            font-family: 'Segoe UI', sans-serif;
            background: #f0f2f5;
            margin: 0;
            padding: 20px;
        }}
        h1 {{
            text-align: center;
            color: #2c3e50;
            margin-bottom: 30px;
        }}
        .query-section {{
            margin-bottom: 40px;
            padding: 20px;
            background: white;
            border-radius: 14px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
        }}
        .row-title {{
            margin-bottom: 16px;
            font-size: 20px;
            font-weight: 600;
            color: #2c3e50;
            border-bottom: 2px solid #eee;
            padding-bottom: 6px;
        }}
        .image-grid {{
            display: flex;
            flex-wrap: wrap;
            gap: 20px;
        }}
        .img-card {{
            background: #fff;
            border-radius: 12px;
            padding: 10px;
            text-align: center;
            width: 280px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
            border: 1px solid #ddd;
        }}
        .img-card:hover {{
            transform: translateY(-4px);
            box-shadow: 0 6px 14px rgba(0,0,0,0.12);
        }}
        .img-card img {{
            width: 100%;
            max-height: 200px;
            object-fit: contain;
            border-radius: 8px;
            border: 2px solid #f1f1f1;
            margin-bottom: 8px;
        }}
        .img-card.query {{
            border: 2px solid #3498db;
        }}
        .meta-info {{
            font-size: 0.85em;
            color: #444;
            text-align: left;
            background: #fafafa;
            padding: 6px;
            border-radius: 6px;
            margin-top: 6px;
        }}
        .meta-info b {{
            color: #2c3e50;
            display: inline-block;
            width: 130px;
        }}
        .sim-score {{
            display: inline-block;
            margin-top: 6px;
            padding: 4px 10px;
            background: #27ae60;
            color: white;
            border-radius: 12px;
            font-size: 0.8em;
        }}
    </style>
</head>
<body>
    <h1>🔍 Image Similarity Viewer</h1>
""")

        for query_idx, neighbors in enumerate(I):
            valid_matches = [
                (idx, D[query_idx][rank])
                for rank, idx in enumerate(neighbors[1:], start=1)
                if (
                    idx < len(valid_paths) and
                    D[query_idx][rank] >= SIMILARITY_THRESHOLD and
                    strip_extension(valid_paths[idx].name) != strip_extension(valid_paths[query_idx].name)
                )
            ]

            if not valid_matches:
                skipped += 1
                continue

            query_path = valid_paths[query_idx]
            meta = metadata_lookup.get(strip_extension(query_path.name), {})
            meta_html = "".join([
                f"<b>{col}:</b> {val}<br>"
                for col, val in meta.items()
                if col not in EXCLUDE_FIELDS and str(val).strip() not in ["nan", "None", ""]
            ]) or "<b>Metadata:</b> Not Found<br>"

            # Section container
            f.write("<div class='query-section'>")
            f.write(f"<div class='row-title'>Query Image: {query_path.name}</div>")
            f.write("<div class='image-grid'>")

            # Query image card
            f.write("<div class='img-card query'>")
            f.write(f"<img src='{IMAGE_DIR}/{query_path.name}' alt='query image'><br>")
            f.write(f"<strong>{query_path.name}</strong>")
            f.write(f"<div class='meta-info'>{meta_html}</div>")
            f.write("<div class='sim-score'>Query Image</div>")
            f.write("</div>")

            # Matched images
            for idx, similarity in valid_matches:
                similar_path = valid_paths[idx]
                similarity_pct = f"{similarity * 100:.1f}%"
                meta = metadata_lookup.get(strip_extension(similar_path.name), {})
                meta_html = "".join([
                    f"<b>{col}:</b> {val}<br>"
                    for col, val in meta.items()
                    if col not in EXCLUDE_FIELDS and str(val).strip() not in ["nan", "None", ""]
                ]) or "<b>Metadata:</b> Not Found<br>"
                f.write("<div class='img-card'>")
                f.write(f"<img src='{IMAGE_DIR}/{similar_path.name}' alt='matched image'><br>")
                f.write(f"<strong>{similar_path.name}</strong>")
                f.write(f"<div class='meta-info'>{meta_html}</div>")
                f.write(f"<div class='sim-score'>Similarity: {similarity_pct}</div>")
                f.write("</div>")

            f.write("</div></div>")  # Close grid and section

        f.write("</body></html>")

    print(f"✅ HTML written to: {OUTPUT_HTML}")
    if skipped > 0:
        print(f"ℹ️ Skipped {skipped} query images with no valid matches above threshold.")

# Helper: Remove file extension
def strip_extension(filename):
    return os.path.splitext(filename)[0]

# Helper: Resize while keeping aspect ratio
def resize_image(image_path, max_width, max_height):
    with Image.open(image_path) as img:
        w, h = img.size
        scale = min(max_width / w, max_height / h)
        return int(w * scale), int(h * scale)
# === EXCEL GENERATOR ===


def strip_extension(filename):
    return os.path.splitext(filename)[0]

def resize_image(image_path, max_width, max_height):
    with Image.open(image_path) as img:
        w, h = img.size
        scale = min(max_width / w, max_height / h)
        return int(w * scale), int(h * scale)




# === EXCEL GENERATOR ===
import os
from openpyxl import Workbook

def strip_extension(filename):
    return os.path.splitext(filename)[0]

def generate_similarity_excel(valid_paths, I, D, OUTPUT_XLSX, SIMILARITY_THRESHOLD):
    print("📑 Generating Excel similarity viewer (Matched Files moved to 2nd column)...")
    skipped = 0

    # Updated headers
    EXCEL_HEADERS = [
        "Query File",
        "Matched Files",
        "Query Design No",
        "Design No & Similarity",
        "Classification",
        "Make",
        "MasterCollection",
        "Collection",
        "Purity",
        "Image URL",
        "ImageName"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Image Similarity"

    # Write headers
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    row = 2

    for query_idx, neighbors in enumerate(I):
        query_path = valid_paths[query_idx]
        query_name = query_path.name

        # Get metadata for the query image
        query_meta = metadata_lookup.get(strip_extension(query_name), {})
        query_design_no = query_meta.get("Design No", strip_extension(query_name))

        # Filter valid matches
        valid_matches = [
            (idx, D[query_idx][rank])
            for rank, idx in enumerate(neighbors[1:], start=1)
            if (
                idx < len(valid_paths)
                and D[query_idx][rank] >= SIMILARITY_THRESHOLD
                and strip_extension(valid_paths[idx].name) != strip_extension(query_name)
            )
        ]

        if not valid_matches:
            skipped += 1
            continue

        # Combine matched files
        matched_files = ", ".join([valid_paths[idx].name for idx, _ in valid_matches])

        # Combine Design No and Similarity pairings
        design_sim_pairs = []
        classification = ""
        make = ""
        master_collection = ""
        collection = ""
        purity = ""
        image_url = ""
        image_name = ""

        for idx, similarity in valid_matches:
            match_meta = metadata_lookup.get(strip_extension(valid_paths[idx].name), {})
            design_no = match_meta.get("Design No", strip_extension(valid_paths[idx].name))
            similarity_pct = f"{similarity*100:.1f}%"
            design_sim_pairs.append(f"{design_no} - {similarity_pct}")

            # Use metadata from first match for other fields
            if not classification:
                classification = match_meta.get("Classification", "")
                make = match_meta.get("Make", "")
                master_collection = match_meta.get("MasterCollection", "")
                collection = match_meta.get("Collection", "")
                purity = match_meta.get("Purity", "")
                image_url = match_meta.get("Image URL", "")
                image_name = match_meta.get("ImageName", "")

        design_sim_text = ", ".join(design_sim_pairs)

        # Write row
        ws.cell(row=row, column=1, value=query_name)
        ws.cell(row=row, column=2, value=matched_files)
        ws.cell(row=row, column=3, value=query_design_no)
        ws.cell(row=row, column=4, value=design_sim_text)
        ws.cell(row=row, column=5, value=classification)
        ws.cell(row=row, column=6, value=make)
        ws.cell(row=row, column=7, value=master_collection)
        ws.cell(row=row, column=8, value=collection)
        ws.cell(row=row, column=9, value=purity)
        ws.cell(row=row, column=10, value=image_url)
        ws.cell(row=row, column=11, value=image_name)

        row += 1

    wb.save(OUTPUT_XLSX)
    print(f"✅ Excel written to: {OUTPUT_XLSX}")
    if skipped > 0:
        print(f"ℹ️ Skipped {skipped} queries with no matches above threshold.")
# ==== RUN VIEWER ====
#generate_similarity_html(valid_paths, I, D, OUTPUT_HTML, SIMILARITY_THRESHOLD)

#html_path = os.path.abspath(OUTPUT_HTML)
#print(f"✅ Done. Opening in browser: {html_path}")
#subprocess.run(["open", html_path])

# ==== RUN VIEWER ====
generate_similarity_excel(valid_paths, I, D, OUTPUT_XLSX, SIMILARITY_THRESHOLD)

excel_path = os.path.abspath(OUTPUT_XLSX)
print(f"✅ Done. Opening in Excel: {excel_path}")
subprocess.run(["open", excel_path])