import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# === IMAGE FILE PATH ===
image_path = "designs/0a0b300c-f9c0-4715-b09f-5893fd4e92c6.jpg"

# Extract only the file name (without extension)
file_name = os.path.splitext(os.path.basename(image_path))[0]

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Image Sheet"

# Write the file name in A1
ws['A1'] = "File Name"
ws['B1'] = "Image"
ws['A2'] = file_name

# Load image
img = Image(image_path)

# Optional: Resize image to fit cell
img.width = 100   # Width in pixels
img.height = 100  # Height in pixels

# Insert image in cell B2
ws.add_image(img, "B2")

# Save workbook
output_file = "image_insert_openpyxl.xlsx"
wb.save(output_file)

print(f"✅ Image '{file_name}' inserted successfully into '{output_file}'")
